from directories import excel_sheet_directory
from openpyxl import load_workbook
from os import path
import csv
import pandas as pd
from matplotlib import pyplot as plt
from scipy import stats
from PhysicsEngine.drawables import Arrow
from scipy.signal import find_peaks
import numpy as np


broken_meters = {'large_20210420115513_20210420120157': [12, 17],
                 'large_20210805171741_20210805172610': [1, 6]} # name of parts, have to subtract one to get index


def find_baseline(signal_data, percentile=5):
    baseline = np.percentile(signal_data, percentile)
    return baseline


def calculate_iqr(signal_data):
    q1 = np.percentile(signal_data, 5)
    q3 = np.percentile(signal_data, 95)
    iqr = q3 - q1
    return iqr


def get_sheet():
    workbook = load_workbook(filename=excel_sheet_directory + path.sep + "Testable.xlsx")
    sheet = workbook.active
    return sheet


sheet = get_sheet()
DISPLAY_CONSTANT = 0.2


class Forces:
    def __init__(self, humans, x):
        self.excel_index = humans.excel_index
        if self.get_force_filename() is not None:
            self.date = self.get_date()
            self.size = humans.size
            self.directory = self.force_directory()
            self.occupied = humans.occupied
            self.filename = self.get_force_filename()
            self.traj_filename = x.filename
            self.abs_values = self.forces_loading(humans.frames, x.fps)
            self.angles = self.get_angles(humans, x)
            self.angles_load = self.angles - x.angle[:, np.newaxis]
            # self.force_meters = Maze(x).force_attachment_positions_in_trajectory(x, reference_frame='load')

    @staticmethod
    def get_angles(humans, x):
        """
        :param humans: object of the class Humans
        :param x: object of the class trajectory_inheritance.trajectory
        :return: angles of the forces in world coordinates
        """
        from trajectory.humans import angle_shift
        angle_shift = np.array([angle_shift[x.size][i]
                                for i in range(len(angle_shift[x.size].keys()))])[np.newaxis, :]
        return humans.angles + x.angle[:, np.newaxis] + angle_shift

    def get_date(self):
        day = sheet.cell(row=self.excel_index, column=2).value
        daytime = sheet.cell(row=self.excel_index, column=6).value
        return day.replace(hour=daytime.hour, minute=daytime.minute)

    def force_directory(self):
        day_string = str(self.date.year) + '-' + str(self.date.month).zfill(2) + '-' + str(self.date.day).zfill(2)
        return ('{0}{1}phys-guru-cs{2}ants{3}Tabea{4}Human Experiments{5}Raw Data and Videos{7}'
                + day_string + '{8}Force Measurements{9}' + self.size).format(path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep)

    def normalize_every_participant(self, q=0.9, cap=1):
        # go through all participants and normalize the force
        for i in range(self.abs_values.shape[1]):
            if np.quantile(self.abs_values[:, i], q) > 1:
                self.abs_values[:, i] = self.abs_values[:, i] / np.quantile(self.abs_values[:, i], q)
            if np.max(self.abs_values[:, i]) > cap:
                self.abs_values[:, i] = np.min([self.abs_values[:, i], np.ones_like(self.abs_values[:, i]) * cap], axis=0)

    def plot_participants(self, frames=(0, -1), parts=range(26)):
        # # plot all self.abs_values
        plt.figure()
        time = np.arange(0, len(self.abs_values))[frames[0]:frames[-1]]
        for i in parts:
            plt.plot(time, self.abs_values[frames[0]:frames[-1], i], label=str(i + 1))
        plt.legend()
        plt.show(block=False)
        plt.xlabel('frames')
        plt.ylabel('force [N]')
        DEBUG = 1

    def adjust_baseline(self, percentile=5):
        # if 20% of the values are below 0, then find the 0.05 percentile and subtract it from the force
        for i in range(self.abs_values.shape[1]):
            if len(np.where(self.abs_values[:, i] < 0)[0]) / self.abs_values.shape[0] > 0.2:
                baseline = find_baseline(self.abs_values[:, i], percentile=percentile)
                self.abs_values[:, i] = self.abs_values[:, i] - baseline

    def remove_extreme_negative_values(self, thresh=-0.3):
        for i in range(self.abs_values.shape[1]):
            # if jumps larger than 10, then smooth over them

            # find ranges which are smaller than -0.5
            def find_ranges(bool_list, min_length=0, max_length=np.inf) -> list:
                inds = np.array([i for i, x in enumerate(bool_list) if x])
                index_successions_in_a = np.split(inds, np.where(np.diff(inds) != 1)[0] + 1)

                ranges = [[ind[0], ind[-1]] for ind in index_successions_in_a if max_length > len(ind) > min_length]
                return ranges
            ranges = find_ranges(self.abs_values[:, i] < thresh)

            for r in ranges:
                if r[0] == 0:
                    mean = np.mean([self.abs_values[r[1]+1, i]], axis=0)
                    self.abs_values[r[0]:r[1]+1, i] = np.ones_like(self.abs_values[r[0]:r[1]+1, i]) * mean
                elif r[1] == len(self.abs_values[:, i]) - 1:
                    mean = np.mean([self.abs_values[r[0]-1, i]], axis=0)
                    self.abs_values[r[0]:r[1]+1, i] = np.ones_like(self.abs_values[r[0]:r[1]+1, i]) * mean
                else:
                    mean = np.mean([self.abs_values[r[0]-1, i], self.abs_values[r[1]+1, i]], axis=0)
                    self.abs_values[r[0]:r[1]+1, i] = np.ones_like(self.abs_values[r[0]:r[1]+1, i]) * mean

    def synchronization_offset(self, fps: int):
        """
        :param fps: frames per second
        If there is no force meter measurement return None.
        :return: frame of turning on force meter relative to start of the raw movie
        """
        if sheet.cell(row=self.excel_index, column=16).value == '/':
            return None

        if sheet.cell(row=self.excel_index, column=16).value is None:
            raise Exception('Fill in the Force synchronization time in line ' + str(self.excel_index))

        [minute, second] = [int(number) for number in
                            sheet.cell(row=self.excel_index, column=16).value.strip()[:-3].split(':')]
        frame_force_meter = (second + minute * 60) * fps

        """ if the frame of synchronization is BEFORE the start of the movie which was tracked """
        if sheet.cell(row=self.excel_index, column=16).value[0] == '-':
            frame_force_meter = - frame_force_meter

        # print(minute, second, self.traj_filename, self.excel_index)

        """ time of tracking relative to start of the raw movie """
        raw_string = sheet.cell(row=self.excel_index, column=8).value
        if ', ' in raw_string:
            frame_tracking = int(raw_string.split(', ')[0])
        else:
            frame_tracking = int(raw_string.split('\n')[0])
        return frame_tracking - frame_force_meter

    def get_force_filename(self):
        txt_name = sheet.cell(row=self.excel_index, column=19).value
        if txt_name.endswith('.txt') or txt_name.endswith('.TXT'):
            return txt_name
        elif txt_name == '/':
            return None
        else:
            raise ValueError('You still have to add the name of the force file in line ' + str(self.excel_index))

    def forces_loading(self, frames, fps):
        from trajectory.humans import participant_number
        # read force meter file
        with open(self.force_directory() + path.sep + self.filename, 'r') as f:
            reader = csv.reader(f, delimiter='\t')
            text_file_content = [line for line in reader]

        def convert_to_frames(fps, times):
            def correct_times():
                # for the large human SPT Maze, we have an additional two digits and an space in our txt file.
                # We have to get rid of this.
                for i, time in enumerate(times):
                    times[i] = [times[i][0].split(' ')[-1]]
                return times

            times = correct_times()
            seconds = [int(time[0].split(':')[0]) * 3600 + int(time[0].split(':')[1]) * 60 + int(time[0].split(':')[2])
                       for time
                       in times]
            seconds = [sec - seconds[0] for sec in seconds]
            frames = []
            for second in range(seconds[-1]):
                measurements_per_second = len([sec for sec in seconds if sec == second])
                for ii in range(measurements_per_second):
                    frames.append(second * fps + int(ii * fps / measurements_per_second))
            return frames

        sampled_frames = convert_to_frames(fps, text_file_content[1::2][1:-1])
        if np.any(np.diff(sampled_frames) > 20) or np.any(np.diff(sampled_frames) < 5):
            print('The force meter is not recording at a constant frame rate', self.traj_filename, self.excel_index)

        forces_txt = [[float(fu) for fu in fo[0].split(' ') if len(fu) > 1] for fo in text_file_content[0::2][:-1]]

        # every frame of the movie gets a force for every force meter
        forces_all_frames = []
        for frames_index in range(len(sampled_frames) - 1):
            for ii in range(sampled_frames[frames_index], sampled_frames[frames_index + 1]):
                forces_all_frames.append(forces_txt[frames_index])

        # find the offset of the first frame of the movie to the start of the force meter measurement
        synch_offset = self.synchronization_offset(fps)

        f = np.array(forces_all_frames)
        baselines = np.array([find_baseline(f[:synch_offset, i], percentile=10) for i in range(f.shape[1])])

        # x = np.arange(0, len(f)) - synch_offset
        # for i in range(f.shape[1]):
        #     plt.plot(x, f[:, i] - baselines[i], label=str(i+1))
        # plt.legend()
        # # draw a line at the synchronization time/fps
        # plt.axvline(x=0, color='r', linestyle='--')
        # plt.ylim(-2, 20)
        # plt.savefig(home + '\\Analysis\\human_alignment\\start\\' + self.traj_filename + '.png', transparent=True, dpi=200)
        # plt.close()

        # write the force into the self.frame[:].forces variable
        if len(forces_all_frames) < len(frames) + synch_offset:
            if sheet.cell(row=self.excel_index, column=18).value is not None and \
                    'battery' in sheet.cell(row=self.excel_index, column=18).value:
                print('Battery empty')
                empty = [0.0 for _ in range(len(forces_all_frames[0]))]
                missing_frames = range(-len(forces_all_frames) + (len(frames) + synch_offset + 10))
                [forces_all_frames.append(empty) for _ in missing_frames]

        abs_values = []
        for i, force_index in enumerate(range(synch_offset, len(frames) + synch_offset)):
            abs_values.append(forces_all_frames[force_index])
        abs_values = np.array(abs_values) - baselines

        # if self.traj_filename in ['large_20210420115513_20210420120157', 'large_20210708001919_20210708003910']:
        #     # I want to ravel axis 1 so that abs_values[:, 0] is abs_values[:, 7]
        shift = 7
        abs_values = np.roll(abs_values, shift, axis=1)
        largest_iqr = np.sort(np.argsort([calculate_iqr(abs_values[:, i])
                                          for i in range(abs_values.shape[1])])[::-1][:len(self.occupied)])

        if np.setdiff1d(self.occupied, largest_iqr).size/len(largest_iqr) > 0.1:
            raise ValueError('The force meters are not in the right order', self.traj_filename, self.excel_index)

        # all unoccupied force meters should have zero force
        empty_indices = [i for i in range(participant_number[self.size]) if i not in self.occupied]
        for empty_index in empty_indices:
            abs_values[:, empty_index] = 0

        if self.traj_filename in broken_meters.keys():
            for i in broken_meters[self.traj_filename]:
                abs_values[:, i-1] = np.nan
        return abs_values

    @staticmethod
    def remove_force_outliers(array):
        def remove_force_outliers_single_forcemeter(single):
            # only one measurement
            df_original = pd.DataFrame(single)

            outlier_index = np.where((np.abs(stats.zscore(df_original, axis=0)) < 5) == False)[0]
            df_original.values[outlier_index] = 0  # TODO: this should be NaN, .. I think
            df_no_outliers = df_original.interpolate()
            return df_no_outliers

        return np.squeeze(np.apply_along_axis(remove_force_outliers_single_forcemeter, 0, array))

    @staticmethod
    def plateaus(arrays):

        def plateau(array):
            plateaus = find_peaks(array, plateau_size=20)[0]
            if len(plateaus) == 0:
                return array.min()
            if len(np.where(array - array[plateaus].mean() < 0)[0]) / len(array) > 0.4:
                return np.nanmin(array)
            return array[plateaus].mean()

        return [plateau(arrays[:, i]) for i in range(arrays.shape[1])]

    def draw(self, display, x):
        force_attachments = display.my_maze.force_attachment_positions()
        for name in x.participants.occupied:
            self.arrow(display.i, force_attachments[name], name).draw(display)

    def torque(self, part):
        return np.cross(self.force_vector(part, reference_frame='load'), self.force_meters[:, part])

    def arrow(self, i, force_meter_coor, name) -> Arrow:
        """
        :param i: index of the participant
        :param force_meter_coor: where is the force_meter located in world coordinates
        :return: start, end and string for the display of the force as a triplet
        """
        start = force_meter_coor
        end = force_meter_coor + self.abs_values[i, name] * DISPLAY_CONSTANT * \
              np.array([np.cos(self.angles[i, name]), np.sin(self.angles[i, name])])
        return Arrow(np.array(start), np.array(end), str(name))

    def force_vector(self, name: int, reference_frame='maze') -> np.ndarray:
        """
        :param name: index of the participant
        :param reference_frame: 'maze' or 'load', dependent on desired reference frame
        :return: len(x.frames)x2 numpy.array with x and y components of the force vectors
        """
        if reference_frame == 'maze':
            a = self.angles[:, name]
        elif reference_frame == 'load':
            a = self.angles_load[:, name]
        else:
            raise ValueError('What frame of reference?')

        return np.transpose(np.array([np.cos(a), np.sin(a)]) * self.abs_values[:, name])
